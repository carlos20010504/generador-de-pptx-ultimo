"""Pure-data Excel parser.

Performs structural-only type inference (no business semantics). The richer
semantic layer lives in `inventory.py`. Dtypes returned here:
  numeric | currency | percent | year | score | bool | date | categorical | text

The structural detectors live in `insights` so they can be re-used and tested
in isolation. If `insights` raises or is unavailable, parser degrades to the
basic pandas-based inference.
"""
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Any
import pandas as pd
from socya_pipeline.errors import PipelineError, ErrorCode
from socya_pipeline import insights

@dataclass
class ColumnData:
    name: str
    dtype: str            # "numeric" | "currency" | "date" | "categorical" | "text" | "bool"
    n_unique: int
    fill_ratio: float
    samples: List[Any] = field(default_factory=list)
    min: Any = None
    max: Any = None
    mean: Any = None
    sum: Any = None
    top_values: List[List[Any]] = field(default_factory=list)  # [[value, count], ...]
    # Unidad inferida del header / cell format (ej. "USD", "kg", "%", "Ton"),
    # None si no se detecta. Usado para etiquetar KPIs y ejes de chart sin
    # tener que adivinar.
    unit: Any = None

@dataclass
class SheetData:
    name: str
    shape: tuple              # (rows, cols)
    fill_ratio: float
    columns: List[ColumnData]
    first_rows: List[List[Any]]   # up to 8 rows in compact form

@dataclass
class WorkbookData:
    filename: str
    sheets: List[SheetData]
    # Named ranges definidos por el usuario en Excel (Insert > Name).
    # Es la señal semántica MÁS fuerte: el usuario etiquetó datos importantes.
    # Lista de {name, sheet, range, scope}. Vacía cuando no hay o falla la lectura.
    named_ranges: List[dict] = field(default_factory=list)

def parse_workbook(path, api_key: str = None) -> WorkbookData:
    """Parse an Excel into a WorkbookData. If `api_key` is provided AND the
    parser detects confusion (no typed columns / many Unnamed headers / etc.),
    a tiny conditional AI inspector runs to suggest dtype overrides. Cached
    aggressively so repeated uploads of the same file don't re-call the AI."""
    p = Path(path)
    try:
        xls = pd.ExcelFile(p, engine=None)  # auto-detect openpyxl/xlrd
    except Exception as e:
        raise PipelineError(
            ErrorCode.EXCEL_INVALID,
            "El archivo no se pudo leer como Excel.",
            details=str(e)[:200],
            user_action="upload_again",
        )

    # Pre-load Excel cell number_formats per (sheet, column). This is the
    # authoritative dtype signal — the user explicitly told Excel "this is
    # currency / percent / date" by applying that format. Falls back silently
    # when openpyxl can't open the file (xlsx-only) or there's no formatting.
    fmt_map = _read_cell_formats(p)

    sheets: List[SheetData] = []
    for sheet_name in xls.sheet_names:
        try:
            df = xls.parse(sheet_name)
        except Exception:
            continue
        df = _resolve_merged_headers(df, p, sheet_name)
        df = _promote_real_headers(xls, sheet_name, df)
        df = _compose_two_row_headers(xls, sheet_name, df)
        df = _strip_total_rows(df)
        df = _filter_formula_errors(df)
        sheets.append(_summarize_sheet(sheet_name, df,
                                         column_formats=fmt_map.get(sheet_name, {})))

    if not sheets:
        raise PipelineError(
            ErrorCode.EXCEL_EMPTY,
            "El Excel se abrió pero no tiene hojas legibles.",
            user_action="upload_again",
        )

    # Named ranges del usuario — best-effort, [] si falla.
    named_ranges = _read_named_ranges(p)

    wb = WorkbookData(filename=p.name, sheets=sheets, named_ranges=named_ranges)

    # Optional AI inspector: ONLY runs when the parser detected confusion AND
    # the caller provided an api_key. Best-effort: any failure is swallowed.
    if api_key:
        try:
            from socya_pipeline.inspector import (
                inspect_workbook, apply_inspector_hints,
            )
            hints = inspect_workbook(wb, p, api_key)
            apply_inspector_hints(wb, hints)
        except Exception:
            pass

    return wb


def promote_real_headers(xls, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """Public wrapper — see _promote_real_headers."""
    return _promote_real_headers(xls, sheet_name, df)


def _promote_real_headers(xls, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """If pandas inferred bad headers (Unnamed:N from a title row, or
    organizer metadata like 'TITLE: ...'), find the real header row and
    re-read from there.

    Triggers:
      - ≥30% of columns are 'Unnamed: N' (typical title-row case), OR
      - The first column header starts with one of the known organizer
        metadata prefixes (TITLE:/SUBTITLE:/TYPE:) — common after the JS
        organizer rewrites the sheet
    """
    cols = [str(c) for c in df.columns]
    unnamed_ratio = sum(1 for c in cols if c.startswith("Unnamed:")) / max(1, len(cols))
    first_col_metadata = (cols and any(cols[0].upper().startswith(p)
                                          for p in ("TITLE:", "SUBTITLE:", "TYPE:")))
    if unnamed_ratio < 0.3 and not first_col_metadata:
        return df

    # Scan first 10 rows for a likely header row
    scan_rows = min(10, len(df))
    n_cols = len(cols)
    # On 1-2 column sheets, can't rely on "≥50% cells filled" — relax to ≥1
    min_required = max(1, int(n_cols * 0.5)) if n_cols > 2 else 1
    for i in range(scan_rows):
        row = df.iloc[i]
        non_null = [v for v in row if pd.notna(v)]
        if len(non_null) < min_required:
            continue
        # Skip organizer metadata rows explicitly
        first_val = str(non_null[0]) if non_null else ""
        if any(first_val.upper().startswith(p)
                for p in ("TITLE:", "SUBTITLE:", "TYPE:")):
            continue
        # A header row has mostly strings, not numbers
        string_cells = sum(1 for v in non_null
                           if isinstance(v, str) and not _looks_numeric(v))
        if string_cells / len(non_null) >= 0.7:
            try:
                return xls.parse(sheet_name, header=i + 1)
            except Exception:
                return df
    return df


def _read_cell_formats(path: Path) -> dict:
    """Use openpyxl to read each sheet's column number_formats. Returns
    {sheet_name: {column_header: format_string}}.

    Strategy: header row is row 1; for each column we sample the format of
    the next data cell (row 2). If row 2 is empty, scan up to row 6.
    Falls back to {} silently when openpyxl can't open the file (e.g. .xls
    legacy format) — the caller will then rely on structural detection."""
    out: dict = {}
    if str(path).lower().endswith((".xls", ".xlsm")):
        # .xls (legacy) is not supported by openpyxl; .xlsm requires keep_vba.
        # For .xlsm we still try; for .xls we skip.
        if str(path).lower().endswith(".xls"):
            return out
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path), read_only=True,
                            data_only=True, keep_links=False)
    except Exception:
        return out

    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
            except Exception:
                continue
            # Iterate first 6 rows to find headers + a sample data cell per col
            rows_iter = ws.iter_rows(min_row=1, max_row=6, values_only=False)
            try:
                rows = list(rows_iter)
            except Exception:
                continue
            if not rows:
                continue
            header_row = rows[0]
            sheet_map: dict = {}
            for col_idx, header_cell in enumerate(header_row):
                header = header_cell.value
                if header is None:
                    continue
                # Find a data cell in the same column with a non-General format
                fmt = None
                for r in rows[1:]:
                    if col_idx >= len(r):
                        continue
                    cell = r[col_idx]
                    if cell is None:
                        continue
                    cf = getattr(cell, "number_format", None)
                    if cf and cf.lower() != "general":
                        fmt = cf
                        break
                if fmt:
                    sheet_map[str(header)] = fmt
            if sheet_map:
                out[sheet_name] = sheet_map
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def strip_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Public wrapper of `_strip_total_rows` — used by the extractor when it
    reloads sheets directly so it gets the same TOTAL-row stripping the parser
    applied during inventory build (otherwise the table slide would re-include
    the TOTAL row even though the KPI sums correctly excluded it)."""
    return _strip_total_rows(df)


def _strip_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop trailing rows that look like totals/subtotals (e.g. last row has
    'TOTAL' in a string cell). These rows otherwise contaminate KPI sums and
    averages by double-counting the data already in the rows above."""
    if df.empty:
        return df
    try:
        # Check up to last 3 rows from the bottom
        candidate_count = 0
        for offset in range(1, min(4, len(df) + 1)):
            row = df.iloc[-offset]
            if insights.is_total_row(row.tolist(), prev_rows_numeric=[]):
                candidate_count += 1
            else:
                break
        if candidate_count > 0:
            return df.iloc[:-candidate_count].reset_index(drop=True)
    except Exception:
        pass
    return df


def _looks_like_date_strings(samples: List) -> bool:
    """Return True if ≥80% of non-null sample values parse as a date.

    pandas often leaves ISO date strings ('2024-01-15') as object dtype
    instead of datetime64; without this check, our parser would call them
    'categorical' and time-series detection would never fire."""
    valid = [s for s in samples if s is not None and str(s).strip()]
    if len(valid) < 3:
        return False
    parsed_ok = 0
    for v in valid:
        try:
            ts = pd.to_datetime(v, errors="raise")
            if pd.notna(ts):
                parsed_ok += 1
        except (ValueError, TypeError, OverflowError):
            continue
    return parsed_ok / len(valid) >= 0.8


def _looks_numeric(s: str) -> bool:
    s = s.strip().replace(",", ".").replace("$", "").replace("%", "")
    try:
        float(s)
        return True
    except (ValueError, AttributeError):
        return False


# Unidades habituales que aparecen en headers entre paréntesis o sufijos.
# Conservador: sólo lo "evidente" para no inventar señales.
_UNIT_RE = __import__("re").compile(
    r"\(\s*("
    r"USD|EUR|COP|MXN|CLP|ARS|BRL|GBP|JPY|CNY|"      # monedas
    r"kg|g|mg|t|ton|lb|"                                # peso
    r"km|m|cm|mm|mi|"                                    # distancia
    r"l|ml|gal|"                                         # volumen
    r"%|pct|"                                            # porcentaje
    r"hab|personas|usuarios|registros|unidades|"        # conteos
    r"hr|h|min|seg|s|días|dias|day|days"                # tiempo
    r")\s*\)",
    flags=__import__("re").IGNORECASE,
)
# Sufijo libre: "Ventas USD" → "USD". Sólo evaluar al FINAL del header.
_UNIT_SUFFIX_RE = __import__("re").compile(
    r"\s+(USD|EUR|COP|MXN|kg|Ton|TON|t|%|hab|km)$"
)


def _detect_unit(header: str, number_format: Any = None) -> Any:
    """Inferir la unidad de la columna a partir del header y/o del cell format.
    Retorna string normalizado (e.g. "USD", "%", "kg") o None."""
    if header:
        h = str(header).strip()
        m = _UNIT_RE.search(h)
        if m:
            raw = m.group(1).strip()
            return _normalize_unit(raw)
        m = _UNIT_SUFFIX_RE.search(h)
        if m:
            return _normalize_unit(m.group(1))
    if number_format:
        nf = str(number_format)
        if "%" in nf:
            return "%"
        # Cell format con currency code embebido: "[$USD-409]" / "[$EUR-407]"
        cur = __import__("re").search(r"\[\$([A-Z]{3})", nf)
        if cur:
            return cur.group(1)
    return None


def _normalize_unit(raw: str) -> str:
    """Normaliza variantes habituales (TON/Ton/t → 't'; pct/% → '%')."""
    s = (raw or "").strip()
    low = s.lower()
    if low in ("ton", "t"):
        return "t"
    if low in ("pct", "porcentaje"):
        return "%"
    if low in ("hab", "habitantes"):
        return "hab"
    if low in ("dias", "días", "day", "days"):
        return "días"
    if low in ("seg", "s"):
        return "s"
    # Currency codes y símbolos las dejamos en mayúscula
    if len(s) == 3 and s.isalpha():
        return s.upper()
    return s


def _read_named_ranges(path: Path) -> List[dict]:
    """Lee defined_names del workbook openpyxl. Cada named range es una
    señal semántica explícita ("Ventas2024", "TotalGeneral", etc.). El AI
    los puede usar para anclar slides a datos importantes para el usuario.
    Best-effort — un workbook sin defined_names devuelve [].
    """
    if not str(path).lower().endswith((".xlsx", ".xlsm")):
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path), read_only=True,
                            data_only=True, keep_links=False)
    except Exception:
        return []
    out: List[dict] = []
    try:
        defined = getattr(wb, "defined_names", None)
        if not defined:
            return out
        # openpyxl 3.x: defined_names is a DefinedNameDict
        try:
            items = defined.items()
        except AttributeError:
            items = []
        for name, dn in items:
            try:
                # dn.value es algo como "Hoja1!$A$1:$B$10"
                value = getattr(dn, "value", None) or ""
                # Skip prints/criterias internos de Excel
                if name.startswith("_xlnm."):
                    continue
                # Parse "SheetName!$A$1:$B$10"
                if "!" in str(value):
                    sheet, rng = str(value).split("!", 1)
                    sheet = sheet.strip("'\"")
                else:
                    sheet, rng = "", str(value)
                out.append({
                    "name": str(name),
                    "sheet": sheet,
                    "range": rng,
                })
            except Exception:
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _compose_two_row_headers(xls, sheet_name: str, df: pd.DataFrame
                                ) -> pd.DataFrame:
    """Detect and collapse a 2-row header into a single composite header.

    Heuristic real-world (informe ejecutivo / financiero):
      Row 0 (df.columns): supercategorías como 'Ventas', 'Costos', 'Margen'
                            con varias columnas seguidas con el mismo valor
                            (o la 1ra con valor + 2 'Unnamed:' a la derecha).
      Row 1 (df.iloc[0]):  encabezados reales por celda — todos strings,
                            mayoría diferentes ('Q1', 'Q2', 'Q3'...).

    Si esto pinta así, componemos `'Ventas > Q1'`, `'Ventas > Q2'`, etc.,
    eliminamos la fila 1 (que pasa a ser header) y devolvemos el df
    re-headerizado. Caso contrario devolvemos df sin cambios.
    """
    if df.empty or len(df) < 2:
        return df

    cols = [str(c) for c in df.columns]
    n = len(cols)
    if n < 2:
        return df

    # Señal A: super-headers se manifiestan en pandas como
    #   (i) 'Unnamed: N' a la derecha del super-header (fila 1 con celdas
    #       vacías a la derecha del valor merged-like); o
    #   (ii) 'Nombre', 'Nombre.1', 'Nombre.2' (pandas auto-suffija
    #        duplicados cuando el super-header NO está merged pero sí
    #        repetido — caso real "Ventas, Ventas, Ventas" en row 1).
    # Cualquiera de las dos señales — ≥25% del total de columnas — sugiere
    # que la fila 0 actual es realmente sub-headers de un super.
    import re as _re_local
    suffix_re = _re_local.compile(r"\.(\d+)$")
    unnamed_after_named = 0
    repeated_after_named = 0
    last_named_root = None
    for c in cols:
        if c.startswith("Unnamed:"):
            if last_named_root:
                unnamed_after_named += 1
            continue
        m = suffix_re.search(c)
        if m and last_named_root:
            # 'Ventas.1' tras 'Ventas' → dedup suffix
            root = c[: m.start()]
            if root == last_named_root:
                repeated_after_named += 1
                continue
        last_named_root = c
    signal = unnamed_after_named + repeated_after_named
    if signal < max(2, int(n * 0.25)):
        return df

    # Señal B: la fila 0 (que sería la 2da fila de headers) es mayormente
    # strings cortos no numéricos, casi todos distintos entre sí.
    try:
        row0 = df.iloc[0].tolist()
    except Exception:
        return df
    str_cells = [v for v in row0 if isinstance(v, str) and v.strip()]
    if len(str_cells) < max(3, int(n * 0.6)):
        return df
    if any(_looks_numeric(s) for s in str_cells):
        return df
    # Diversidad mínima — si todos son iguales no es un sub-header útil
    uniq_ratio = len(set(s.strip().lower() for s in str_cells)) / max(1, len(str_cells))
    if uniq_ratio < 0.6:
        return df

    # Componer headers: forward-fill el supercategoría sobre los Unnamed
    # y los duplicados con dedup-suffix de pandas (Ventas, Ventas.1, etc.)
    super_filled: List[str] = []
    last = ""
    for c in cols:
        # Normalizar: quitar dedup suffix '.N' y NaN para inferir root
        m = suffix_re.search(c)
        root = c[: m.start()] if m else c
        if not root.startswith("Unnamed:") and root.lower() not in ("nan", ""):
            last = root
        super_filled.append(last)

    new_headers: List[str] = []
    for i, super_h in enumerate(super_filled):
        sub = ""
        try:
            v = row0[i]
            if v is not None and str(v).strip():
                sub = str(v).strip()
        except Exception:
            pass
        if super_h and sub:
            new_headers.append(f"{super_h} > {sub}")
        elif super_h:
            new_headers.append(super_h)
        elif sub:
            new_headers.append(sub)
        else:
            new_headers.append(f"Col {i + 1}")

    new_df = df.iloc[1:].reset_index(drop=True).copy()
    new_df.columns = new_headers
    return new_df


def _resolve_merged_headers(df: pd.DataFrame, path: Path, sheet_name: str,
                              header_row: int = 1) -> pd.DataFrame:
    """If the sheet has merged cells in/around the header row, replicate the
    top-left value into the right neighbor headers. Two real-world variants:

      A) Merged super-header SOLO en row 1 (no hay sub-headers en row 2):
         'Total' merged C1:E1 → df.columns = ['Total', 'Unnamed:3', 'Unnamed:4']
         → componemos como ['Total', 'Total_2', 'Total_3'].

      B) Merged super-header en row 1 + sub-headers en row 2:
         Row 1 'Q1 2024' merged B1:D1, 'Q2 2024' merged E1:G1.
         Row 2: Ventas, Costos, Margen, Ventas, Costos, Margen.
         → componemos df.columns como ['Q1 2024 > Ventas', 'Q1 2024 > Costos', ...]
         y CONSUMIMOS row 0 (que era row 2 del Excel) — pasa a ser fila de
         datos, no de sub-headers. Sin esto los sub-headers se contaminarían
         con el dtype de las columnas.

    Detección B vs A: si la PRIMERA fila de datos (df.iloc[0]) tiene mostly
    strings no-numéricos en las columnas Unnamed, asumimos sub-headers.
    """
    if not str(path).lower().endswith((".xlsx", ".xlsm")):
        return df
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path), read_only=False, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return df
        ws = wb[sheet_name]
        ranges = list(ws.merged_cells.ranges) if ws.merged_cells else []
        if not ranges:
            wb.close()
            return df

        # Build a map of {col_idx_0based: super_header_value} from merged
        # ranges that intersect the header row.
        super_map: dict = {}
        for mrange in ranges:
            if mrange.min_row > header_row or mrange.max_row < header_row:
                continue
            top_left = ws.cell(row=mrange.min_row, column=mrange.min_col).value
            if top_left is None or not str(top_left).strip():
                continue
            for col in range(mrange.min_col, mrange.max_col + 1):
                super_map[col - 1] = str(top_left).strip()
        wb.close()

        if not super_map:
            return df

        # Detectar variante B: ¿la primera fila de datos en las columnas
        # afectadas tiene strings (sub-headers) o números (data)?
        cur_cols = list(df.columns)
        is_variant_b = False
        if len(df) > 0:
            try:
                row0 = df.iloc[0].tolist()
                affected_indices = [i for i in super_map if 0 <= i < len(row0)]
                if affected_indices:
                    str_count = 0
                    for i in affected_indices:
                        v = row0[i]
                        if (isinstance(v, str) and v.strip()
                            and not _looks_numeric(v)):
                            str_count += 1
                    if str_count >= max(2, len(affected_indices) * 0.6):
                        is_variant_b = True
            except Exception:
                pass

        new_cols = list(cur_cols)
        if is_variant_b:
            # Componer "Super > Sub" usando row 0 como sub-headers
            row0 = df.iloc[0].tolist()
            for i in range(len(new_cols)):
                if i in super_map:
                    sub = ""
                    if i < len(row0):
                        v = row0[i]
                        if v is not None and str(v).strip():
                            sub = str(v).strip()
                    if sub:
                        new_cols[i] = f"{super_map[i]} > {sub}"
                    else:
                        new_cols[i] = super_map[i]
            # Consumimos row 0 (sub-headers no son datos)
            df = df.iloc[1:].reset_index(drop=True).copy()
            df.columns = new_cols
        else:
            # Variante A: forward-fill de super sobre Unnamed con suffix
            for i in range(len(new_cols)):
                if i in super_map:
                    current = str(new_cols[i])
                    if current.startswith("Unnamed:") or current.lower() == "nan":
                        # Find la posición relativa dentro del rango merged
                        # contando cuántos hermanos del mismo super están
                        # antes de este idx.
                        same_super_before = sum(
                            1 for j in range(i) if super_map.get(j) == super_map[i]
                        )
                        suffix = same_super_before + 1
                        new_cols[i] = (f"{super_map[i]}_{suffix}" if suffix > 1
                                         else super_map[i])
            if new_cols != cur_cols:
                df = df.copy()
                df.columns = new_cols
    except Exception:
        # Best-effort — never fail the parse over merged cell resolution.
        pass
    return df


# Excel formula errors that bleed into pandas as strings (e.g. "#DIV/0!").
# We strip them so they don't end up rendered as values in slides or
# polluting numeric aggregations.
_FORMULA_ERROR_TOKENS = (
    "#N/A", "#DIV/0!", "#REF!", "#NAME?", "#VALUE!",
    "#NUM!", "#NULL!", "#GETTING_DATA", "#SPILL!", "#CALC!",
)


def _filter_formula_errors(df: pd.DataFrame) -> pd.DataFrame:
    """Replace cells whose value is an Excel formula error (e.g. "#DIV/0!",
    "#N/A") with NaN. Pandas leaves them as strings, which makes them sneak
    into KPI labels and numeric series as garbage.
    """
    if df.empty:
        return df
    try:
        # Vectorised replace across object/string columns. Numeric columns
        # already won't contain these tokens. (Pandas 3 will require both
        # 'object' AND 'str' to be explicit; we pass them now to avoid the
        # deprecation warning when the user upgrades.)
        obj_cols = df.select_dtypes(include=["object", "str"]).columns
        if len(obj_cols) == 0:
            return df

        def _strip(v):
            if isinstance(v, str):
                v_up = v.strip().upper()
                if v_up in _FORMULA_ERROR_TOKENS or v_up.startswith("#"):
                    # Be conservative: only strip if it really looks like a
                    # formula error token (starts with '#' AND ends with '!'
                    # or equals one of the known tokens).
                    if v_up in _FORMULA_ERROR_TOKENS or v_up.endswith(("!", "?")):
                        return None
            return v
        for c in obj_cols:
            df[c] = df[c].apply(_strip)
    except Exception:
        pass
    return df


def _summarize_sheet(name: str, df: pd.DataFrame,
                       column_formats: dict = None) -> SheetData:
    column_formats = column_formats or {}
    rows, cols = df.shape
    if rows == 0 or cols == 0:
        return SheetData(name=name, shape=(rows, cols), fill_ratio=0.0,
                         columns=[], first_rows=[])
    fill_ratio = float(df.notna().sum().sum()) / max(1, rows * cols)
    columns = [_summarize_column(c, df[c],
                                    excel_format=column_formats.get(str(c)))
                for c in df.columns]
    first_rows = df.head(8).where(df.head(8).notna(), None).values.tolist()
    return SheetData(name=name, shape=(rows, cols), fill_ratio=fill_ratio,
                     columns=columns, first_rows=first_rows)


def _summarize_column(name: Any, series: pd.Series,
                        excel_format: str = None) -> ColumnData:
    name_str = str(name)
    fill = float(series.notna().sum()) / max(1, len(series))
    n_unique = int(series.nunique(dropna=True))
    dtype = _infer_dtype(name_str, series, excel_format=excel_format)
    samples = series.dropna().head(8).tolist()
    unit = _detect_unit(name_str, number_format=excel_format)

    col = ColumnData(name=name_str, dtype=dtype, n_unique=n_unique,
                     fill_ratio=fill, samples=_jsonify(samples), unit=unit)

    if dtype in ("numeric", "currency", "percent", "year", "score"):
        nums = pd.to_numeric(series, errors="coerce")
        # Replace ±inf with NaN so they don't poison the aggregations
        nums = nums.replace([float("inf"), float("-inf")], pd.NA).dropna()
        if not nums.empty:
            def _safe(v):
                v = float(v)
                return v if math.isfinite(v) else None
            col.min  = _safe(nums.min())
            col.max  = _safe(nums.max())
            col.mean = _safe(nums.mean())
            col.sum  = _safe(nums.sum())
    elif dtype in ("categorical", "text"):
        vc = series.dropna().astype(str).value_counts().head(8)
        col.top_values = [[v, int(c)] for v, c in vc.items()]

    return col


def _infer_dtype(name: str, series: pd.Series,
                  excel_format: str = None) -> str:
    """Structural type inference. Priority order:
    1. Excel cell number_format (the user's explicit declaration — authoritative)
    2. Pandas dtype (bool, datetime)
    3. Structural detectors (year/score/percent/currency-by-magnitude)
    4. Name-based hints (i18n money tokens)
    5. Cardinality (categorical vs text)
    """
    # 1. Authoritative: the user's Excel cell formatting
    fmt_dtype = None
    try:
        fmt_dtype = insights.format_to_dtype(excel_format)
    except Exception:
        pass
    if fmt_dtype == "currency":
        return "currency"
    if fmt_dtype == "percent":
        return "percent"
    if fmt_dtype == "date":
        return "date"
    # fmt_dtype == "numeric" or None → keep going (need more info to refine)

    if pd.api.types.is_bool_dtype(series):
        return "bool"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"

    samples_for_struct = series.dropna().head(80).tolist()

    if pd.api.types.is_numeric_dtype(series):
        # Year columns (1900-2100 integers) — must precede currency/numeric
        try:
            if insights.is_year_column(samples_for_struct, name=name):
                return "year"
        except Exception:
            pass
        # Score / rating (small range integers)
        try:
            if insights.is_score_column(samples_for_struct, name):
                return "score"
        except Exception:
            pass
        # Percentage (either 0..1 always, or 0..100 with name hint)
        try:
            if insights.is_percentage_column(samples_for_struct, name):
                return "percent"
        except Exception:
            pass
        # Currency by name (i18n list) takes priority — explicit signal
        try:
            if insights.looks_money_by_name(name):
                return "currency"
        except Exception:
            pass
        # Currency by structural magnitude (no name needed)
        try:
            if insights.is_currency_by_magnitude(samples_for_struct):
                return "currency"
        except Exception:
            pass
        return "numeric"

    # Non-numeric: try date coercion FIRST (catches ISO strings like
    # "2024-01-15" that pandas left as object dtype)
    try:
        if _looks_like_date_strings(samples_for_struct):
            return "date"
    except Exception:
        pass

    # Then check for boolean-disguised string columns
    try:
        if insights.looks_boolean_disguised(samples_for_struct):
            return "bool"
    except Exception:
        pass

    n_unique = series.nunique(dropna=True)
    n_non_null = int(series.notna().sum())

    # Reject categorical when:
    #   - every non-null is unique (it's a free-text or ID column, not a
    #     finite set of categories), OR
    #   - sample values are long (>25 chars on average) → looks like prose
    if n_non_null > 0 and n_unique == n_non_null and n_non_null >= 4:
        return "text"
    if samples_for_struct:
        avg_len = sum(len(str(s)) for s in samples_for_struct
                       if s is not None) / max(1, len(samples_for_struct))
        if avg_len > 25:
            return "text"

    if n_unique > 0 and n_unique <= max(20, len(series) * 0.05):
        return "categorical"
    return "text"


def _jsonify(values):
    out = []
    for v in values:
        if pd.isna(v):
            out.append(None)
        elif hasattr(v, "isoformat"):
            out.append(v.isoformat())
        elif isinstance(v, (int, float, str, bool)):
            out.append(v)
        else:
            out.append(str(v))
    return out
