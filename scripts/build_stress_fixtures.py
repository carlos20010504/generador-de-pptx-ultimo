"""Generate a battery of stress-testing Excel fixtures.

Each fixture exercises a different real-world failure mode to make sure the
pipeline (parser → inventory → onboarding → planner → extractor → renderer)
behaves robustly across the full input space — not just on clean Excels.

Output: __tests__/stress_fixtures/*.xlsx

Run: python scripts/build_stress_fixtures.py
"""
from __future__ import annotations

import os
import random
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

random.seed(42)
np.random.seed(42)

OUT_DIR = Path("__tests__/stress_fixtures")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def _save(wb: Workbook, name: str) -> Path:
    p = OUT_DIR / name
    wb.save(str(p))
    print(f"  [OK] {name}  ({p.stat().st_size // 1024} KB)")
    return p


# ─────────────────────────────────────────────────────────────────────
# Fixture 1: tiny_minimal — el caso degenerado
# ─────────────────────────────────────────────────────────────────────

def fx_tiny_minimal() -> Path:
    """3 filas, 2 cols. Justo el mínimo para producir KPIs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Producto", "Ventas"])
    ws.append(["A", 100])
    ws.append(["B", 250])
    ws.append(["C", 175])
    return _save(wb, "01_tiny_minimal.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 2: clean_simple — caso real "ideal"
# ─────────────────────────────────────────────────────────────────────

def fx_clean_simple() -> Path:
    """100 filas, 6 cols, headers limpios, dtypes claros."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comisiones"
    ws.append(["Fecha", "Solicitante", "Ciudad", "Estado",
                "Monto Solicitado", "Monto Aprobado"])
    cities = ["Bogotá", "Medellín", "Cali", "Barranquilla", "Cartagena"]
    states = ["APROBADO", "RECHAZADO", "EN REVISIÓN"]
    state_weights = [0.65, 0.20, 0.15]
    base = datetime(2024, 1, 1)
    for i in range(100):
        d = base + timedelta(days=i * 3)
        sol = random.choice(states)
        m_sol = random.randint(50_000, 5_000_000)
        m_apr = m_sol if sol == "APROBADO" else (
            int(m_sol * 0.7) if sol == "EN REVISIÓN" else 0
        )
        ws.append([d, f"Solicitante {i + 1:03d}",
                   random.choice(cities),
                   random.choices(states, weights=state_weights)[0],
                   m_sol, m_apr])
    # Aplicar formato de moneda a Monto Solicitado y Aprobado
    for col_letter in ("E", "F"):
        for row in range(2, 102):
            ws[f"{col_letter}{row}"].number_format = '"$"#,##0'
    return _save(wb, "02_clean_simple.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 3: multi_sheet — relaciones entre hojas
# ─────────────────────────────────────────────────────────────────────

def fx_multi_sheet() -> Path:
    """4 hojas: Maestro + 3 hojas categóricas que se relacionan por id."""
    wb = Workbook()
    ws_master = wb.active
    ws_master.title = "Maestro"
    ws_master.append(["id_cliente", "Nombre", "Segmento", "Ingreso Anual"])
    segments = ["Premium", "Estándar", "Básico"]
    for i in range(50):
        ws_master.append([
            f"C{i + 1:04d}",
            f"Cliente {i + 1}",
            random.choice(segments),
            random.randint(20_000_000, 200_000_000),
        ])
    for r in range(2, 52):
        ws_master[f"D{r}"].number_format = '"$"#,##0'

    ws_tx = wb.create_sheet("Transacciones")
    ws_tx.append(["id_cliente", "Fecha", "Tipo", "Monto"])
    base = datetime(2024, 1, 1)
    for i in range(300):
        ws_tx.append([
            f"C{random.randint(1, 50):04d}",
            base + timedelta(days=random.randint(0, 365)),
            random.choice(["Compra", "Devolución", "Suscripción"]),
            round(random.uniform(50_000, 2_000_000), 2),
        ])
    for r in range(2, 302):
        ws_tx[f"D{r}"].number_format = '"$"#,##0.00'

    ws_geo = wb.create_sheet("Geografía")
    ws_geo.append(["Departamento", "Municipio", "Población"])
    deps = [("Antioquia", "Medellín", 2_500_000),
            ("Antioquia", "Envigado", 230_000),
            ("Cundinamarca", "Bogotá D.C.", 7_800_000),
            ("Valle del Cauca", "Cali", 2_300_000),
            ("Atlántico", "Barranquilla", 1_200_000),
            ("Bolívar", "Cartagena", 980_000)]
    for d in deps:
        ws_geo.append(d)
    for r in range(2, len(deps) + 2):
        ws_geo[f"C{r}"].number_format = "#,##0"

    ws_hist = wb.create_sheet("Histórico Mensual")
    ws_hist.append(["Mes", "Ventas", "Costos", "Utilidad"])
    for m in range(1, 25):
        d = datetime(2023, 1, 1) + timedelta(days=30 * (m - 1))
        v = random.randint(40_000_000, 90_000_000)
        c = int(v * random.uniform(0.5, 0.75))
        ws_hist.append([d.strftime("%Y-%m"), v, c, v - c])
    for col in ("B", "C", "D"):
        for r in range(2, 26):
            ws_hist[f"{col}{r}"].number_format = '"$"#,##0'

    # Named range para señalizar dato importante
    dn = DefinedName(name="VentasYTD", attr_text="'Histórico Mensual'!$B$2:$B$25")
    wb.defined_names["VentasYTD"] = dn

    return _save(wb, "03_multi_sheet.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 4: merged_headers — celdas combinadas
# ─────────────────────────────────────────────────────────────────────

def fx_merged_headers() -> Path:
    """Headers con merged cells: super-categoría que abarca varias columnas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Q1-Q4"
    # Row 1: super-headers (con merge)
    ws.append(["Categoría", "Q1 2024", None, None,
                "Q2 2024", None, None, "Total"])
    # Row 2: sub-headers reales
    ws.append(["", "Ventas", "Costos", "Margen",
                "Ventas", "Costos", "Margen", ""])
    # Merged ranges en row 1
    ws.merge_cells("B1:D1")  # Q1 2024
    ws.merge_cells("E1:G1")  # Q2 2024
    # Data
    cats = ["Producto A", "Producto B", "Producto C", "Producto D",
            "Servicio X", "Servicio Y"]
    for cat in cats:
        q1v = random.randint(50_000_000, 200_000_000)
        q1c = int(q1v * 0.6)
        q2v = random.randint(50_000_000, 200_000_000)
        q2c = int(q2v * 0.6)
        ws.append([cat, q1v, q1c, q1v - q1c, q2v, q2c, q2v - q2c,
                   q1v + q2v])
    for col in ("B", "C", "D", "E", "F", "G", "H"):
        for r in range(3, 9):
            ws[f"{col}{r}"].number_format = '"$"#,##0'
    return _save(wb, "04_merged_headers.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 5: multi_row_headers — encabezados de 2 filas
# ─────────────────────────────────────────────────────────────────────

def fx_multi_row_headers() -> Path:
    """Two-row composite headers: super-categoría arriba, sub-encabezados
    distintos por columna abajo. SIN merged cells (sólo estructura visual)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"
    # Row 1: super (forward-filled estilo informe)
    ws.append(["", "Ventas", "Ventas", "Ventas", "Costos", "Costos", "Margen"])
    # Row 2: detalle
    ws.append(["Mes", "Producto A", "Producto B", "Producto C",
                "Operativos", "Logística", "Total"])
    base = datetime(2024, 1, 1)
    for m in range(12):
        ws.append([
            (base + timedelta(days=30 * m)).strftime("%Y-%m"),
            random.randint(10_000_000, 30_000_000),
            random.randint(8_000_000, 25_000_000),
            random.randint(5_000_000, 20_000_000),
            random.randint(15_000_000, 35_000_000),
            random.randint(5_000_000, 12_000_000),
            random.randint(5_000_000, 30_000_000),
        ])
    for col in "BCDEFG":
        for r in range(3, 15):
            ws[f"{col}{r}"].number_format = '"$"#,##0'
    return _save(wb, "05_multi_row_headers.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 6: formula_errors — #N/A, #DIV/0!, #REF!
# ─────────────────────────────────────────────────────────────────────

def fx_formula_errors() -> Path:
    """Mix de valores reales con errores de fórmula como strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos con errores"
    ws.append(["Cliente", "Solicitado", "Aprobado", "Tasa Aprobación"])
    errors = ["#N/A", "#DIV/0!", "#REF!", "#NAME?", "#VALUE!"]
    for i in range(40):
        sol = random.randint(100_000, 5_000_000)
        # 20% de las filas tienen error en aprobado o tasa
        if random.random() < 0.20:
            apr = random.choice(errors)
            tasa = random.choice(errors)
        else:
            apr = int(sol * random.uniform(0.4, 1.0))
            tasa = round(apr / sol, 4) if sol else 0
        ws.append([f"Cliente {i + 1}", sol, apr, tasa])
    # Format columnas numéricas
    for r in range(2, 42):
        ws[f"B{r}"].number_format = '"$"#,##0'
        # No format Apr ni Tasa porque algunos son strings
    return _save(wb, "06_formula_errors.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 7: messy_unicode — caracteres "raros" que rompen Windows cp1252
# ─────────────────────────────────────────────────────────────────────

def fx_messy_unicode() -> Path:
    """Emojis, narrow space ( ), em-dashes, diacríticos compuestos.
    Esto sí o sí rompía el pipeline en Windows antes del fix UTF-8."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aliados estratégicos"
    ws.append(["Aliado", "País — Región", "Valoración", "Notas"])
    rows = [
        ("Acueducto Sur 💧", "Colombia — Pacífico", 4.5, "Líder en cobertura"),
        ("Energía Verde 🌱", "Brasil — Nordeste", 4.2, "Crecimiento 2024"),
        ("Telcom Pro 📡", "México — Norte", 3.8, "Expansión rural"),
        ("Banco Andino 🏦", "Perú — Lima", 4.0, "Microcrédito"),
        ("EduTech ⚙️", "Argentina – Patagonia", 4.7, "Plataforma EdTech"),
        ("AguaSegura 🚰", "Ecuador — Sierra", 3.5, "Saneamiento básico"),
    ]
    for r in rows:
        ws.append(r)
    return _save(wb, "07_messy_unicode.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 8: large_dataset — volumen
# ─────────────────────────────────────────────────────────────────────

def fx_large_dataset() -> Path:
    """10K filas, 12 columnas. Para verificar performance + memoria."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Operaciones"
    ws.append([
        "id_op", "Fecha", "Tipo", "Subtipo", "Cliente",
        "Departamento", "Municipio", "Producto",
        "Cantidad", "Precio Unitario", "Subtotal", "IVA",
    ])
    deps = [
        ("Antioquia", "Medellín"), ("Antioquia", "Envigado"),
        ("Cundinamarca", "Bogotá D.C."), ("Cundinamarca", "Soacha"),
        ("Valle del Cauca", "Cali"), ("Valle del Cauca", "Palmira"),
        ("Atlántico", "Barranquilla"), ("Bolívar", "Cartagena"),
    ]
    products = ["Servicio Básico", "Servicio Premium", "Consultoría",
                "Implementación", "Soporte", "Capacitación"]
    types = ["B2B", "B2C", "B2G"]
    subtypes = ["Recurrente", "Ad-hoc", "Proyecto"]
    base = datetime(2022, 1, 1)
    for i in range(10_000):
        dep, mun = random.choice(deps)
        qty = random.randint(1, 20)
        unit = round(random.uniform(50_000, 500_000), 2)
        sub = qty * unit
        iva = round(sub * 0.19, 2)
        ws.append([
            f"OP{i + 1:06d}",
            base + timedelta(days=random.randint(0, 730)),
            random.choice(types),
            random.choice(subtypes),
            f"Cliente {random.randint(1, 500):03d}",
            dep, mun,
            random.choice(products),
            qty, unit, sub, iva,
        ])
    for col in ("J", "K", "L"):
        for r in range(2, 10_002):
            ws[f"{col}{r}"].number_format = '"$"#,##0.00'
    return _save(wb, "08_large_dataset.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 9: pii_sensitive — emails, cédulas, teléfonos
# ─────────────────────────────────────────────────────────────────────

def fx_pii_sensitive() -> Path:
    """Datos personales evidentes — para verificar PII detection."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiarios"
    ws.append(["Cédula", "Nombre", "Email", "Teléfono",
                "Ciudad", "Subsidio"])
    first_names = ["María", "Juan", "Ana", "Carlos", "Laura", "Diego",
                    "Sofía", "Andrés", "Camila", "Felipe"]
    last_names = ["García", "Rodríguez", "Martínez", "López", "Pérez",
                   "González", "Hernández", "Torres", "Ramírez", "Castro"]
    cities = ["Bogotá", "Medellín", "Cali", "Bucaramanga", "Pereira"]
    for i in range(60):
        nombre = f"{random.choice(first_names)} {random.choice(last_names)}"
        cedula = f"{random.randint(10_000_000, 1_500_000_000)}"
        email = f"{nombre.lower().replace(' ', '.')}@correo.com"
        tel = f"+57 3{random.randint(10, 25)} {random.randint(100, 999)} {random.randint(1000, 9999)}"
        ws.append([
            cedula, nombre, email, tel,
            random.choice(cities),
            random.randint(500_000, 3_000_000),
        ])
    for r in range(2, 62):
        ws[f"F{r}"].number_format = '"$"#,##0'
    return _save(wb, "09_pii_sensitive.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 10: time_series_long — daily data, 2 años
# ─────────────────────────────────────────────────────────────────────

def fx_time_series_long() -> Path:
    """730 días de data diaria. Pone a prueba aggregation y forecasting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Diario"
    ws.append(["Fecha", "Visitas", "Conversiones", "Ingreso (USD)"])
    base = datetime(2023, 1, 1)
    # Tendencia + seasonality + ruido
    for d in range(730):
        date = base + timedelta(days=d)
        # Trend creciente
        trend = 1000 + d * 1.5
        # Seasonality semanal (más weekdays)
        weekday_boost = 1.3 if date.weekday() < 5 else 0.7
        # Ruido
        noise = random.uniform(0.85, 1.15)
        visits = int(trend * weekday_boost * noise)
        conv = int(visits * random.uniform(0.02, 0.06))
        ingreso = round(conv * random.uniform(50, 200), 2)
        ws.append([date, visits, conv, ingreso])
    for r in range(2, 732):
        ws[f"D{r}"].number_format = '"$"#,##0.00'
    return _save(wb, "10_time_series_long.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 11: pathological — casos degenerados
# ─────────────────────────────────────────────────────────────────────

def fx_pathological() -> Path:
    """Casos límite que NO deberían romper:
    - Hoja vacía
    - Hoja con solo título
    - Columna toda NaN
    - Columna con un solo valor único
    - Fila TOTAL al final
    - Mix de tipos en una columna
    """
    wb = Workbook()
    # Hoja 1: vacía
    ws_empty = wb.active
    ws_empty.title = "Vacía"
    # No append nada

    # Hoja 2: solo título (no es tabla)
    ws_title = wb.create_sheet("Solo título")
    ws_title.append(["Reporte Anual de Gestión 2024"])

    # Hoja 3: tabla con TOTAL al final + columna toda NaN + columna constante
    ws_tot = wb.create_sheet("Datos parciales")
    ws_tot.append(["Categoría", "Valor", "Observación", "Año"])
    ws_tot.append(["A", 100, None, 2024])
    ws_tot.append(["B", 200, None, 2024])
    ws_tot.append(["C", 300, None, 2024])
    ws_tot.append(["D", 150, None, 2024])
    ws_tot.append(["TOTAL", 750, None, 2024])

    # Hoja 4: mix de tipos en columna "Cantidad" (números + strings)
    ws_mix = wb.create_sheet("Mixed types")
    ws_mix.append(["Producto", "Cantidad", "Estado"])
    rows_mix = [
        ("A", 100, "OK"),
        ("B", "N/D", "Pendiente"),
        ("C", 250, "OK"),
        ("D", "—", "Sin info"),
        ("E", 175.5, "OK"),
        ("F", 0, "Cero"),
    ]
    for r in rows_mix:
        ws_mix.append(r)

    return _save(wb, "11_pathological.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Fixture 12: with_units — unidades en headers
# ─────────────────────────────────────────────────────────────────────

def fx_with_units() -> Path:
    """Headers con unidades explícitas — verifica detección automática."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores ESG"
    ws.append([
        "Proyecto",
        "Energía consumida (kWh)",
        "Emisiones (Ton CO2)",
        "Beneficiarios (hab)",
        "Inversión (USD)",
        "Eficiencia (%)",
    ])
    projects = ["Solar Antioquia", "Eólico Guajira", "Hídrico Santander",
                "Reforestación Amazonas", "Educación rural Cauca",
                "Agua potable Chocó", "Saneamiento Nariño"]
    for p in projects:
        ws.append([
            p,
            random.randint(50_000, 500_000),
            round(random.uniform(50, 800), 1),
            random.randint(500, 50_000),
            random.randint(100_000, 5_000_000),
            round(random.uniform(60, 95), 1),
        ])
    for r in range(2, 9):
        ws[f"E{r}"].number_format = '"$"#,##0'
        ws[f"F{r}"].number_format = "0.0"
    return _save(wb, "12_with_units.xlsx")


# ─────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────

def main():
    print(f"Generando fixtures en {OUT_DIR.resolve()}\n")
    fixtures = [
        fx_tiny_minimal,
        fx_clean_simple,
        fx_multi_sheet,
        fx_merged_headers,
        fx_multi_row_headers,
        fx_formula_errors,
        fx_messy_unicode,
        fx_large_dataset,
        fx_pii_sensitive,
        fx_time_series_long,
        fx_pathological,
        fx_with_units,
    ]
    failures = []
    for fn in fixtures:
        try:
            fn()
        except Exception as e:
            print(f"  [FAIL] {fn.__name__}: {e}")
            failures.append(fn.__name__)
    print(f"\n{len(fixtures) - len(failures)}/{len(fixtures)} fixtures generados.")
    return 1 if failures else 0


if __name__ == "__main__":
    import sys
    sys.exit(main() or 0)
